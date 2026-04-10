"""
batch_processing.py
-------------------
Reads date_last_updated_*.xlsx, divides the ~1 000 active sites into
ordered batches of 150 (most recently updated first), writes "yes" into
the aw_active column of the target Google Sheet for each batch, and
saves a local batch_log_*.xlsx tracking every assignment.

Batch logic
-----------
  Batch 1 : all sites updated within the past 1 month  (if ≤ 150 sites)
             OR the 150 most recently updated sites     (if > 150 in that window)
  Batch 2+ : the next 150 sites in recency order, exclusive of prior batches
  … repeat until all ~1 000 sites are assigned.

The script is safe to run incrementally: it reads batch_log_*.xlsx
(if one exists) so already-assigned sites are never repeated.

Usage
-----
    python batch_processing.py                   # process next un-run batch
    python batch_processing.py --all             # process all batches in one go
    python batch_processing.py --dry-run         # preview only, no Sheet writes
    python batch_processing.py --batch 3         # run a specific batch number

Prerequisites
-------------
    pip install gspread google-auth pandas openpyxl

Google Sheets authentication (choose ONE):

  Option A – Service Account (recommended for automation):
    1. Go to console.cloud.google.com → APIs & Services → Credentials
    2. Create a Service Account, download the JSON key
    3. Share the Google Sheet with the service-account email (Editor access)
    4. Set:  CREDENTIALS_FILE = "path/to/service_account.json"

  Option B – Personal OAuth (simpler for one-off use):
    1. pip install gspread[oauth]
    2. Run once interactively; gspread will open a browser for consent
    3. Set:  CREDENTIALS_FILE = None   (triggers OAuth flow)
"""

import glob
import sys
import argparse
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
SHEET_ID         = "14mfVh7zL-uuzkqBbf2EwgNs07Lj4E9n2WKY727c0CYI"
WORKSHEET_GID    = 2107301382       # the specific tab (gid= in the URL)
BATCH_SIZE       = 150
INITIAL_WINDOW   = 30               # days – lookback for "first month" logic
CREDENTIALS_FILE = None             # path to service_account.json, or None for OAuth

# ── Authentication ────────────────────────────────────────────────────────────

def _gs_client():
    try:
        import gspread
    except ImportError:
        raise SystemExit(
            "[batch] gspread not installed.\n"
            "        Run:  pip install gspread google-auth"
        )
    if CREDENTIALS_FILE:
        return gspread.service_account(filename=CREDENTIALS_FILE)
    else:
        return gspread.oauth()


def _open_worksheet():
    import gspread as _gs
    gc = _gs_client()
    sh = gc.open_by_key(SHEET_ID)
    ws = next((w for w in sh.worksheets() if w.id == WORKSHEET_GID), None)
    if ws is None:
        raise RuntimeError(
            f"Worksheet with gid={WORKSHEET_GID} not found in spreadsheet {SHEET_ID}.\n"
            "Check that the tab is still there and you have Editor access."
        )
    return ws


# ── Input file helpers ────────────────────────────────────────────────────────

def _find_dates_file() -> str:
    candidates = sorted(glob.glob("date_last_updated_*.xlsx"), reverse=True)
    if not candidates:
        raise FileNotFoundError(
            "No date_last_updated_*.xlsx found.\n"
            "Run:  python get_last_updated.py  first."
        )
    chosen = candidates[0]
    print(f"[batch] Using dates file: {chosen}")
    return chosen


def _find_batch_log() -> str | None:
    candidates = sorted(glob.glob("batch_log_*.xlsx"), reverse=True)
    return candidates[0] if candidates else None


# ── Batch assignment logic ────────────────────────────────────────────────────

def _load_already_assigned(log_path: str | None) -> set[str]:
    """Return the set of URLs already assigned to any batch."""
    if log_path is None:
        return set()
    try:
        df = pd.read_excel(log_path)
        return set(df["url"].dropna().str.strip())
    except Exception:
        return set()


def build_all_batches(df_dates: pd.DataFrame, already_assigned: set[str]) -> list[pd.DataFrame]:
    """
    Sort sites by date_last_updated (newest first) and split into
    successive batches of BATCH_SIZE, excluding already-assigned URLs.

    Batch 1 special rule:
      - If ≤ BATCH_SIZE sites were updated within INITIAL_WINDOW days,
        take all of them as batch 1.
      - Otherwise take the BATCH_SIZE most recent.
    All subsequent batches are straight slices of BATCH_SIZE.
    """
    # Drop sites with no date and exclude already-assigned
    df_valid = df_dates.dropna(subset=["date_last_updated"]).copy()
    df_valid = df_valid[~df_valid["url"].str.strip().isin(already_assigned)]
    df_valid = df_valid.sort_values("date_last_updated", ascending=False).reset_index(drop=True)

    if df_valid.empty:
        return []

    batches: list[pd.DataFrame] = []
    cutoff  = date.today() - timedelta(days=INITIAL_WINDOW)
    pos     = 0

    # First-batch special rule
    within_month = df_valid[
        df_valid["date_last_updated"].apply(
            lambda d: d >= cutoff if isinstance(d, date) else False
        )
    ]
    if 0 < len(within_month) <= BATCH_SIZE:
        # Fewer than BATCH_SIZE sites updated recently: take all of them
        b1  = within_month
        pos = len(within_month)
    else:
        # Either >BATCH_SIZE recent sites, or none at all: take top BATCH_SIZE
        b1  = df_valid.iloc[:BATCH_SIZE]
        pos = BATCH_SIZE

    if not b1.empty:
        batches.append(b1.copy())

    # Remaining batches: straight slices
    while pos < len(df_valid):
        chunk = df_valid.iloc[pos : pos + BATCH_SIZE]
        if chunk.empty:
            break
        batches.append(chunk.copy())
        pos += BATCH_SIZE

    return batches


# ── Google Sheet update ───────────────────────────────────────────────────────

def update_google_sheet(batch_df: pd.DataFrame, batch_num: int, dry_run: bool) -> int:
    """
    Write "yes" into the aw_active column for each site in batch_df.
    Matches rows by URL. Returns the number of cells updated.
    """
    if dry_run:
        print(f"  [dry-run] Would update {len(batch_df)} rows in Google Sheet.")
        return len(batch_df)

    ws        = _open_worksheet()
    all_vals  = ws.get_all_values()
    if not all_vals:
        print("  [warn] Google Sheet is empty.")
        return 0

    headers   = [h.strip().lower() for h in all_vals[0]]
    try:
        url_col = headers.index("url")
        aw_col  = headers.index("aw_active")
    except ValueError as e:
        raise RuntimeError(f"Expected column not found in Google Sheet: {e}")

    # Build url → row-number map (1-based; row 1 = header)
    url_to_row: dict[str, int] = {}
    for i, row in enumerate(all_vals[1:], start=2):
        if url_col < len(row):
            url_to_row[row[url_col].strip()] = i

    batch_urls = set(batch_df["url"].str.strip())
    updates    = []
    unmatched  = []

    for url in batch_urls:
        if url in url_to_row:
            r   = url_to_row[url]
            import gspread as _gs
            col = _gs.utils.rowcol_to_a1(r, aw_col + 1)  # gspread is 1-based
            updates.append({"range": col, "values": [["yes"]]})
        else:
            unmatched.append(url)

    if updates:
        ws.batch_update(updates)

    if unmatched:
        print(f"  [warn] {len(unmatched)} URLs from batch {batch_num} not found in Sheet:")
        for u in unmatched[:5]:
            print(f"         {u}")
        if len(unmatched) > 5:
            print(f"         … and {len(unmatched)-5} more")

    matched = len(updates)
    print(f"  [sheet] Updated {matched} rows in Google Sheet (batch {batch_num}).")
    return matched


# ── Batch log writer ──────────────────────────────────────────────────────────

def _append_to_log(batches_run: list[tuple[int, pd.DataFrame]], dry_run: bool) -> str:
    """
    Writes (or appends to) batch_log_YYYY-MM-DD.xlsx recording every
    site assigned to a batch, with columns:
        batch_number | name | state | url | date_last_updated | asset_count | assigned_on
    """
    log_path = f"batch_log_{date.today()}.xlsx"

    rows = []
    for batch_num, df in batches_run:
        for _, r in df.iterrows():
            rows.append({
                "batch_number":      batch_num,
                "name":              r.get("name", ""),
                "state":             r.get("state", ""),
                "url":               r.get("url", ""),
                "date_last_updated": r.get("date_last_updated", ""),
                "asset_count":       r.get("asset_count", ""),
                "assigned_on":       str(date.today()),
                "dry_run":           dry_run,
            })

    # If log already exists from today, load and append
    existing_rows: list[dict] = []
    if Path(log_path).exists():
        try:
            existing_df  = pd.read_excel(log_path)
            existing_rows = existing_df.to_dict("records")
        except Exception:
            pass

    all_rows = existing_rows + rows
    out_df   = pd.DataFrame(all_rows)

    FONT    = "Arial"
    HDR_CLR = "2E4057"

    def _font(bold=False, color="000000"):
        return Font(name=FONT, bold=bold, size=10, color=color)
    def _fill(h):
        return PatternFill("solid", fgColor=h)
    def _border():
        t = Side(style="thin", color="CCCCCC")
        return Border(left=t, right=t, top=t, bottom=t)

    BATCH_COLORS = [
        "D6F0D6", "D6E8FA", "FFF3CC", "F0D6FA",
        "FAE6D6", "D6F0F0", "FAD6D6", "E8FAD6",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Log"

    cols       = list(out_df.columns)
    col_widths = [14, 28, 8, 55, 18, 12, 12, 10]

    # Header
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=h.replace("_", " ").title())
        cell.font      = _font(bold=True, color="FFFFFF")
        cell.fill      = _fill(HDR_CLR)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _border()

    # Data
    for r_i, record in enumerate(all_rows, start=2):
        bnum  = int(record.get("batch_number", 1))
        color = BATCH_COLORS[(bnum - 1) % len(BATCH_COLORS)]
        fill  = _fill(color)
        for c_i, col in enumerate(cols, 1):
            val  = record.get(col, "")
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.font      = _font()
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = _border()
            if col == "date_last_updated" and isinstance(val, date):
                cell.number_format = "YYYY-MM-DD"

    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    wb.save(log_path)
    return log_path


# ── Summary printer ───────────────────────────────────────────────────────────

def _print_batch_summary(batch_num: int, df: pd.DataFrame) -> None:
    oldest = df["date_last_updated"].min()
    newest = df["date_last_updated"].max()
    today  = date.today()
    days_ago_oldest = (today - oldest).days if isinstance(oldest, date) else "?"
    days_ago_newest = (today - newest).days if isinstance(newest, date) else "?"
    print(
        f"\n  Batch {batch_num}  ({len(df)} sites)\n"
        f"  Most recent update : {newest}  ({days_ago_newest} days ago)\n"
        f"  Oldest update      : {oldest}  ({days_ago_oldest} days ago)\n"
    )


# ── Main ──────────────────────────────────────────────────────────────────────

def run(
    target_batch: int | None = None,
    run_all: bool            = False,
    dry_run: bool            = False,
) -> None:
    print(f"[batch] dry_run={dry_run}  target_batch={target_batch}  run_all={run_all}\n")

    dates_file  = _find_dates_file()
    df_dates    = pd.read_excel(dates_file)

    # Coerce date column
    df_dates["date_last_updated"] = pd.to_datetime(
        df_dates["date_last_updated"], errors="coerce"
    ).dt.date

    log_path    = _find_batch_log()
    already     = _load_already_assigned(log_path)
    print(f"[batch] {len(already)} sites already assigned in prior batches.\n")

    all_batches = build_all_batches(df_dates, already)
    if not all_batches:
        print("[batch] No remaining sites to assign. All done!")
        return

    total_sites = sum(len(b) for b in all_batches)
    print(f"[batch] {len(all_batches)} batches to run  ({total_sites} remaining sites)\n")

    # Determine which batches to process this run
    # Offset batch numbers by how many already existed
    existing_batch_count = _count_existing_batches(log_path)
    numbered = [(existing_batch_count + i + 1, df) for i, df in enumerate(all_batches)]

    if target_batch is not None:
        to_run = [(n, df) for n, df in numbered if n == target_batch]
        if not to_run:
            print(f"[batch] Batch {target_batch} not found in remaining work.")
            return
    elif run_all:
        to_run = numbered
    else:
        to_run = [numbered[0]]   # default: just the next batch

    batches_done: list[tuple[int, pd.DataFrame]] = []

    for batch_num, batch_df in to_run:
        _print_batch_summary(batch_num, batch_df)

        print(f"  Sites in batch {batch_num}:")
        for _, r in batch_df.iterrows():
            dt = r["date_last_updated"]
            print(f"    {str(r['name'])[:35]:<35}  last updated: {dt}")

        updated = update_google_sheet(batch_df, batch_num, dry_run=dry_run)
        batches_done.append((batch_num, batch_df))
        print(f"  ✓ Batch {batch_num} complete — {updated} Sheet rows updated.\n")

    log_out = _append_to_log(batches_done, dry_run=dry_run)
    print(f"[batch] Batch log saved → '{log_out}'\n")


def _count_existing_batches(log_path: str | None) -> int:
    if log_path is None:
        return 0
    try:
        df = pd.read_excel(log_path)
        return int(df["batch_number"].max()) if "batch_number" in df.columns else 0
    except Exception:
        return 0


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Batch-assign active sites and update Google Sheets.")
    parser.add_argument("--all",      action="store_true", help="Run all remaining batches at once.")
    parser.add_argument("--dry-run",  action="store_true", help="Preview batches without writing to Google Sheets.")
    parser.add_argument("--batch",    type=int, default=None, metavar="N", help="Run only batch N.")
    args = parser.parse_args()

    run(
        target_batch = args.batch,
        run_all      = args.all,
        dry_run      = args.dry_run,
    )
