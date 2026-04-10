"""
get_last_updated.py
-------------------
For every site that returned at least one asset in the 1-year scraper run,
find the date of its most recently uploaded document and write a ranked
spreadsheet sorted newest → oldest.

This is the required input for batch_processing.py.

Usage:
    python get_last_updated.py                           # auto-finds newest scraper_results_*.xlsx
    python get_last_updated.py scraper_results_2026-02-23.xlsx

Prerequisites:
    pip install civic-scraper pandas openpyxl requests
"""

import glob
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
LOOKBACK_DAYS = 365        # must match the window used in verify_civic_scraper.py
MAX_WORKERS   = 10         # concurrent civic-scraper calls

# ── Helpers ───────────────────────────────────────────────────────────────────

def _find_results_file(explicit: str | None) -> str:
    if explicit:
        return explicit
    candidates = sorted(glob.glob("scraper_results_*.xlsx"), reverse=True)
    if not candidates:
        raise FileNotFoundError(
            "No scraper_results_*.xlsx found in the current directory.\n"
            "Run:  python verify_civic_scraper.py scraper  first."
        )
    chosen = candidates[0]
    print(f"[dates] Using results file: {chosen}")
    return chosen


def _date_window() -> tuple[str, str]:
    end   = date.today()
    start = end - timedelta(days=LOOKBACK_DAYS)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def _coerce_date(val) -> date | None:
    """Try to extract a plain date from whatever type civic-scraper returns."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S", "%B %d, %Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _latest_date_for_asset(asset) -> date | None:
    """Return the best available date from an Asset object."""
    for attr in ("meeting_date", "start_date", "date", "scraped_at", "created_at"):
        val = getattr(asset, attr, None)
        d = _coerce_date(val)
        if d is not None:
            return d
    return None


# ── Per-site scrape ───────────────────────────────────────────────────────────

def get_latest_date(row: pd.Series) -> tuple[date | None, int]:
    """
    Run civic-scraper on a single site and return (most_recent_date, asset_count).
    Returns (None, 0) on any error.
    """
    url   = str(row.get("url", "")).strip()
    stype = str(row.get("site_type", "")).strip().lower()

    start_dt, end_dt = _date_window()

    try:
        if stype == "civicplus":
            from civic_scraper.platforms import CivicPlusSite
            site   = CivicPlusSite(url)
            assets = site.scrape(start_date=start_dt, end_date=end_dt)
        elif stype == "legistar":
            from civic_scraper.platforms import LegistarSite
            site   = LegistarSite(url)
            assets = site.scrape(start_date=start_dt, end_date=end_dt)
        else:
            return None, 0

        if not assets:
            return None, 0

        dates = [d for a in assets if (d := _latest_date_for_asset(a)) is not None]
        count = len(assets)
        return (max(dates) if dates else None), count

    except Exception:
        return None, 0


def _task(idx: int, row: pd.Series) -> tuple[int, date | None, int]:
    latest, count = get_latest_date(row)
    return idx, latest, count


# ── Main ──────────────────────────────────────────────────────────────────────

def run(results_file: str | None = None) -> str:
    """
    Reads the scraper_results file, scrapes each active site for its most
    recent asset date, and saves date_last_updated_YYYY-MM-DD.xlsx.

    Returns the path of the written file.
    """
    try:
        import civic_scraper  # noqa: F401
    except ImportError:
        raise SystemExit(
            "[dates] ERROR: civic-scraper not installed.\n"
            "        Run:  pip install civic-scraper"
        )

    src = _find_results_file(results_file)
    # The scraper_results file has 3 metadata rows before the real header (row 4).
    # Column display names use title case + spaces ("Assets Found"), so we
    # normalise them to snake_case to match the rest of this script.
    df_src = pd.read_excel(src, header=3)
    df_src.columns = [
        c.lower().replace(" ", "_").replace("(", "").replace(")", "")
        for c in df_src.columns
    ]

    # Keep only sites where civic-scraper found at least one asset
    active = df_src[df_src["assets_found"] == True].copy().reset_index(drop=True)  # noqa: E712
    n = len(active)
    print(f"[dates] {n} active sites to date-check\n")

    latest_dates: dict[int, date | None] = {}
    asset_counts: dict[int, int]         = {}

    t0 = time.time()
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_task, i, row): i for i, row in active.iterrows()}
        done = 0
        for fut in as_completed(futures):
            idx, latest, count = fut.result()
            latest_dates[idx] = latest
            asset_counts[idx] = count
            done += 1
            icon = "✓" if latest else "✗"
            dt_str = str(latest) if latest else "no date found"
            name = str(active.at[idx, "name"])[:40]
            print(f"  {icon} [{done:4d}/{n}]  {name:<40}  {dt_str}  ({count} assets)")

    elapsed = time.time() - t0
    found = sum(1 for d in latest_dates.values() if d is not None)
    print(f"\n[dates] Done – {found}/{n} sites had datable assets  ({elapsed:.1f}s)\n")

    # Build output dataframe
    KEEP = ["name", "state", "country", "gov_level", "site_type", "url"]
    out = active[KEEP].copy()
    out["date_last_updated"] = [latest_dates.get(i) for i in out.index]
    out["asset_count"]       = [asset_counts.get(i, 0) for i in out.index]

    # Sort newest first; undated sites go to the bottom
    out = out.sort_values("date_last_updated", ascending=False, na_position="last")
    out = out.reset_index(drop=True)

    out_path = f"date_last_updated_{date.today()}.xlsx"
    _write_report(out, out_path)
    print(f"[dates] Saved → '{out_path}'\n")
    return out_path


# ── Excel writer ──────────────────────────────────────────────────────────────

def _write_report(df: pd.DataFrame, path: str) -> None:
    FONT     = "Arial"
    HDR_CLR  = "2E4057"
    ALT_CLR  = "F4F8FB"
    DATE_CLR = "D6F0D6"   # green for datable rows
    NONE_CLR = "FAD7D7"   # pink for rows with no date

    def _font(bold=False, size=10, color="000000"):
        return Font(name=FONT, bold=bold, size=size, color=color)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border():
        t = Side(style="thin", color="CCCCCC")
        return Border(left=t, right=t, top=t, bottom=t)

    wb = Workbook()
    ws = wb.active
    ws.title = "Date Last Updated"
    ws.freeze_panes = "A3"

    HEADERS    = ["Name", "State", "Country", "Gov Level", "Site Type",
                  "URL", "Date Last Updated", "Asset Count"]
    COL_WIDTHS = [28, 8, 10, 14, 12, 55, 18, 13]

    # — Title row —
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    t = ws.cell(row=1, column=1,
                value=f"Most Recent Asset Date by Site  ·  Lookback: {LOOKBACK_DAYS} days  ·  {date.today()}")
    t.font      = _font(bold=True, size=11)
    t.fill      = _fill("F0F4F8")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    # — Header row —
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font      = _font(bold=True, color="FFFFFF")
        cell.fill      = _fill(HDR_CLR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
    ws.row_dimensions[2].height = 16

    # — Data rows —
    for r, (_, row) in enumerate(df.iterrows(), start=3):
        has_date = pd.notna(row["date_last_updated"]) and row["date_last_updated"] is not None
        fill     = _fill(DATE_CLR if has_date else NONE_CLR)
        vals     = [
            row["name"], row["state"], row["country"], row["gov_level"],
            row["site_type"], row["url"],
            row["date_last_updated"] if has_date else "—",
            int(row["asset_count"]) if pd.notna(row["asset_count"]) else 0,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = _font()
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = _border()
            # Format date column
            if c == 7 and has_date:
                cell.number_format = "YYYY-MM-DD"

    for c, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(path)


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    src_file = sys.argv[1] if len(sys.argv) > 1 else None
    run(src_file)
