#!/usr/bin/env python3
"""
merge_fips_back.py
------------------
Copies fully-enriched gov_level, state_fips, and county_fips values from:
    Copy of AW_civic_scraper_sites_updated.xlsx   (source — already enriched)
into:
    Copy of AW_civic_scraper_sites.xlsx            (target — original file)

The merge is performed row-by-row on matching DataFrame index positions.
Only cells that are NULL / NaN in the target are overwritten; cells that
already have a value are left untouched (idempotent).

After writing, a validation report is printed confirming 0 remaining nulls
for active US rows across all three geographic columns.

Usage:
    /Users/kateli/Desktop/BigLocal/venv/bin/python3 \
        /Users/kateli/Desktop/BigLocal/merge_fips_back.py
"""

import os
import sys
import pandas as pd
from copy import copy
import openpyxl
from openpyxl import load_workbook

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE   = "/Users/kateli/Desktop/BigLocal/Original Sites to Verify"
ORIG   = os.path.join(BASE, "Copy of AW_civic_scraper_sites.xlsx")
UPDT   = os.path.join(BASE, "Copy of AW_civic_scraper_sites_updated.xlsx")

GEO_COLS = ["gov_level", "state_fips", "county_fips"]

US_STATE_ABBRS = {
    'AL','AK','AZ','AR','CA','CO','CT','DE','FL','GA',
    'HI','ID','IL','IN','IA','KS','KY','LA','ME','MD',
    'MA','MI','MN','MS','MO','MT','NE','NV','NH','NJ',
    'NM','NY','NC','ND','OH','OK','OR','PA','RI','SC',
    'SD','TN','TX','UT','VT','VA','WA','WV','WI','WY','DC'
}

# ── Step 1: Load both files with pandas ───────────────────────────────────────
print("Loading files …")
orig_df = pd.read_excel(ORIG, dtype=str)
updt_df = pd.read_excel(UPDT, dtype=str)

print(f"  Original : {len(orig_df)} rows, columns: {list(orig_df.columns)}")
print(f"  Updated  : {len(updt_df)} rows, columns: {list(updt_df.columns)}")

# ── Step 2: Sanity checks ──────────────────────────────────────────────────────
if len(orig_df) != len(updt_df):
    sys.exit(
        f"ERROR: Row count mismatch — original has {len(orig_df)} rows, "
        f"updated has {len(updt_df)} rows. Aborting."
    )

missing_in_updt = [c for c in GEO_COLS if c not in updt_df.columns]
if missing_in_updt:
    sys.exit(f"ERROR: Updated file is missing columns: {missing_in_updt}. Aborting.")

missing_in_orig = [c for c in GEO_COLS if c not in orig_df.columns]
if missing_in_orig:
    print(f"NOTE: Original file is missing columns {missing_in_orig} — they will be added.")

print("  Sanity checks passed.")

# ── Step 3: Audit nulls BEFORE merge ──────────────────────────────────────────
# Identify active US rows in the original
def is_active_us(row):
    active = str(row.get('aw_active', '')).strip().lower()
    state  = str(row.get('state', '')).strip().upper()
    return active in ('true', '1', 'yes') and state in US_STATE_ABBRS

active_mask = orig_df.apply(is_active_us, axis=1)
active_count = active_mask.sum()
print(f"\nActive US rows in original: {active_count}")

print("\nNull counts BEFORE merge (active US rows only):")
for col in GEO_COLS:
    if col in orig_df.columns:
        null_n = orig_df.loc[active_mask, col].isna().sum()
        # Also treat empty strings as null
        empty_n = (orig_df.loc[active_mask, col].fillna('').str.strip() == '').sum()
        print(f"  {col}: {null_n} NaN  ({empty_n} blank/NaN combined)")
    else:
        print(f"  {col}: COLUMN MISSING (will be added)")

# ── Step 4: Load the target workbook with openpyxl to preserve formatting ─────
print("\nLoading original workbook with openpyxl (to preserve formatting) …")
wb = load_workbook(ORIG)
ws = wb.active

# Build a map: column_name -> 1-based column index in the worksheet
header_row = [cell.value for cell in ws[1]]
col_index = {name: idx + 1 for idx, name in enumerate(header_row) if name is not None}
print(f"  Worksheet columns found: {list(col_index.keys())}")

# Ensure all three geo columns exist in the worksheet; add them if not
for col_name in GEO_COLS:
    if col_name not in col_index:
        new_col_idx = ws.max_column + 1
        ws.cell(row=1, column=new_col_idx, value=col_name)
        col_index[col_name] = new_col_idx
        print(f"  Added missing column '{col_name}' at position {new_col_idx}")

# ── Step 5: Merge values row by row ───────────────────────────────────────────
print("\nMerging enriched values into original workbook …")

cells_written = {col: 0 for col in GEO_COLS}
cells_skipped_existing = {col: 0 for col in GEO_COLS}
cells_skipped_no_source = {col: 0 for col in GEO_COLS}

# pandas rows are 0-indexed; worksheet data rows start at row 2 (row 1 = header)
for pandas_idx in range(len(orig_df)):
    ws_row = pandas_idx + 2  # openpyxl row number

    for col_name in GEO_COLS:
        # Current value in original
        orig_cell = ws.cell(row=ws_row, column=col_index[col_name])
        orig_val  = orig_cell.value

        # Skip if already populated (idempotent)
        if orig_val is not None and str(orig_val).strip() not in ('', 'nan', 'None'):
            cells_skipped_existing[col_name] += 1
            continue

        # Get the enriched value from the updated dataframe
        if col_name not in updt_df.columns:
            cells_skipped_no_source[col_name] += 1
            continue

        new_val = updt_df.at[pandas_idx, col_name]

        # Treat pandas NaN / 'nan' / '' as "no value available"
        if pd.isna(new_val) or str(new_val).strip() in ('', 'nan', 'None', '<NA>'):
            cells_skipped_no_source[col_name] += 1
            continue

        # For FIPS columns: store as integer (strip leading zeros — Excel will
        # show the raw number; zero-padding is enforced at query/export time).
        # gov_level stays as string.
        if col_name in ('state_fips', 'county_fips'):
            try:
                orig_cell.value = int(float(str(new_val).strip()))
            except (ValueError, TypeError):
                orig_cell.value = str(new_val).strip()
        else:
            orig_cell.value = str(new_val).strip()

        cells_written[col_name] += 1

print("\nMerge stats:")
for col_name in GEO_COLS:
    print(f"  {col_name}:")
    print(f"    Written (was null, now filled) : {cells_written[col_name]}")
    print(f"    Skipped (already had a value)  : {cells_skipped_existing[col_name]}")
    print(f"    Skipped (no source value)      : {cells_skipped_no_source[col_name]}")

# ── Step 6: Save the original file in place ───────────────────────────────────
print(f"\nSaving enriched workbook to: {ORIG}")
wb.save(ORIG)
print("  Saved successfully.")

# ── Step 7: Post-merge validation ─────────────────────────────────────────────
print("\n--- POST-MERGE VALIDATION ---")
final_df = pd.read_excel(ORIG, dtype=str)
final_active_mask = final_df.apply(is_active_us, axis=1)
final_active_count = final_active_mask.sum()
print(f"Active US rows after merge: {final_active_count}")

all_clear = True
print("\nNull counts AFTER merge (active US rows only):")
for col in GEO_COLS:
    if col in final_df.columns:
        combined_null = (
            final_df.loc[final_active_mask, col]
            .fillna('')
            .str.strip()
            .isin(['', 'nan', 'None'])
            .sum()
        )
        status = "OK" if combined_null == 0 else "STILL HAS GAPS"
        if combined_null > 0:
            all_clear = False
        print(f"  {col}: {combined_null} null/blank  [{status}]")
    else:
        print(f"  {col}: COLUMN STILL MISSING  [ERROR]")
        all_clear = False

print()
if all_clear:
    print("VALIDATION PASSED: All active US rows have gov_level, state_fips, and county_fips populated.")
else:
    print("VALIDATION WARNING: Some gaps remain. Review the counts above.")
    # Print a sample of still-missing rows for diagnosis
    for col in GEO_COLS:
        if col not in final_df.columns:
            continue
        still_null = (
            final_df.loc[final_active_mask, col]
            .fillna('')
            .str.strip()
            .isin(['', 'nan', 'None'])
        )
        if still_null.sum() > 0:
            print(f"\n  Sample rows still missing '{col}':")
            sample = final_df.loc[final_active_mask & still_null, ['name', 'state', 'gov_level', 'state_fips', 'county_fips']].head(10)
            print(sample.to_string())

print("\nDone.")
