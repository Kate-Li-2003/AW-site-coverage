#!/usr/bin/env python3
"""
legistar_get_last_updated.py
-----------------------------
For every Legistar site that returned assets in the recent 2-month scrape,
re-scrape with a 1-year lookback window to find the most recently uploaded
document date.  Output is a formatted spreadsheet sorted newest → oldest,
analogous to AW_civic_scraper_sites2026-03-02.xlsx.

Usage:
    python3.9 legistar_get_last_updated.py
    python3.9 legistar_get_last_updated.py --input Original\ Sites\ to\ Verify/legistar_scrape_results_2026-04-01.xlsx
    python3.9 legistar_get_last_updated.py --workers 8

Output:
    Original Sites to Verify/legistar_date_last_updated_<date>.xlsx
"""

import argparse
import glob
import time
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Config ────────────────────────────────────────────────────────────────────
BASE          = "Original Sites to Verify"
LOOKBACK_DAYS = 365
MAX_WORKERS   = 8
TODAY         = date.today()
OUTPUT_FILE   = f"{BASE}/legistar_date_last_updated_{TODAY}.xlsx"

# ── Date window ───────────────────────────────────────────────────────────────
def _date_window():
    end   = TODAY
    start = end - timedelta(days=LOOKBACK_DAYS)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


# ── Date coercion ─────────────────────────────────────────────────────────────
def _coerce_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S", "%B %d, %Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _latest_date_for_asset(asset):
    for attr in ("meeting_date", "start_date", "date", "scraped_at", "created_at"):
        val = getattr(asset, attr, None)
        d = _coerce_date(val)
        if d is not None:
            return d
    return None


# ── Per-site scrape ───────────────────────────────────────────────────────────
def get_latest_date(url: str):
    """
    Scrape one Legistar site over the 1-year window.
    Returns (most_recent_date | None, asset_count).
    """
    from civic_scraper.platforms import LegistarSite

    start_dt, end_dt = _date_window()
    try:
        site   = LegistarSite(url, timezone="US/Eastern")
        assets = site.scrape(start_date=start_dt, end_date=end_dt)
        if not assets:
            return None, 0
        dates = [d for a in assets if (d := _latest_date_for_asset(a)) is not None]
        return (max(dates) if dates else None), len(assets)
    except Exception as exc:
        return None, 0  # silently drop errors — already catalogued in scrape results


def _task(idx, row):
    latest, count = get_latest_date(row["url_used"])
    return idx, latest, count


# ── Find latest results file ──────────────────────────────────────────────────
def _find_results_file(explicit):
    if explicit:
        return explicit
    candidates = sorted(
        glob.glob(f"{BASE}/legistar_scrape_results_*.xlsx"), reverse=True
    )
    if not candidates:
        raise FileNotFoundError(
            f"No legistar_scrape_results_*.xlsx found in {BASE}/\n"
            "Run legistar_scraper_all_sites.py first."
        )
    chosen = candidates[0]
    print(f"Using results file: {chosen}")
    return chosen


# ── Main ──────────────────────────────────────────────────────────────────────
def run(results_file=None, workers=MAX_WORKERS):
    src = _find_results_file(results_file)
    df_src = pd.read_excel(src, dtype=str)

    # Keep only sites where the previous scrape succeeded (returned assets)
    successes = df_src[df_src["scrape_status"] == "success"].copy().reset_index(drop=True)
    n = len(successes)
    start_dt, end_dt = _date_window()

    print("=" * 65)
    print("LEGISTAR — DATE LAST UPDATED")
    print("=" * 65)
    print(f"  Source file  : {src}")
    print(f"  Sites        : {n} (scrape_status == 'success')")
    print(f"  Date window  : {start_dt} → {end_dt}  ({LOOKBACK_DAYS} days)")
    print(f"  Workers      : {workers}")
    print()

    latest_dates = {}
    asset_counts = {}

    t0 = time.time()
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_task, i, row): i for i, row in successes.iterrows()}
        done = 0
        for fut in as_completed(futures):
            idx, latest, count = fut.result()
            latest_dates[idx] = latest
            asset_counts[idx] = count
            done += 1
            icon   = "✓" if latest else "○"
            dt_str = str(latest) if latest else "no date found"
            name   = str(successes.at[idx, "name"])[:40]
            state  = str(successes.at[idx, "state"])
            print(f"  {icon} [{done:3d}/{n}]  [{state}] {name:<38}  {dt_str}  ({count} assets)")

    elapsed = time.time() - t0
    found   = sum(1 for d in latest_dates.values() if d is not None)
    print(f"\nDone — {found}/{n} sites had datable assets  ({elapsed:.0f}s)\n")

    # Build output dataframe
    out = successes[["name", "state", "site_type", "aw_active", "endpoint", "url_used"]].copy()
    out["date_last_updated"] = [latest_dates.get(i) for i in out.index]
    out["asset_count_1yr"]   = [asset_counts.get(i, 0) for i in out.index]

    # Sort newest first; undated at bottom
    out = out.sort_values("date_last_updated", ascending=False, na_position="last")
    out = out.reset_index(drop=True)

    _write_report(out, OUTPUT_FILE)
    print(f"Saved → {OUTPUT_FILE}")
    return OUTPUT_FILE


# ── Excel writer ──────────────────────────────────────────────────────────────
def _write_report(df, path):
    FONT     = "Arial"
    HDR_CLR  = "2E4057"
    DATE_CLR = "D6F0D6"   # green  — datable rows
    NONE_CLR = "FAD7D7"   # pink   — no date found

    def _font(bold=False, size=10, color="000000"):
        return Font(name=FONT, bold=bold, size=size, color=color)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border():
        t = Side(style="thin", color="CCCCCC")
        return Border(left=t, right=t, top=t, bottom=t)

    wb = Workbook()
    ws = wb.active
    ws.title = "Date Last Updated"
    ws.freeze_panes = "A3"

    HEADERS    = ["Name", "State", "Site Type", "AW Active",
                  "Endpoint", "URL Used", "Date Last Updated", "Asset Count (1yr)"]
    COL_WIDTHS = [30, 8, 14, 10, 40, 48, 20, 18]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    t = ws.cell(row=1, column=1,
                value=f"Legistar — Most Recent Asset Date  ·  Lookback: {LOOKBACK_DAYS} days  ·  {TODAY}")
    t.font      = _font(bold=True, size=11)
    t.fill      = _fill("F0F4F8")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18

    # Header row
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font      = _font(bold=True, color="FFFFFF")
        cell.fill      = _fill(HDR_CLR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
    ws.row_dimensions[2].height = 16

    # Data rows
    for r, (_, row) in enumerate(df.iterrows(), start=3):
        has_date = pd.notna(row["date_last_updated"]) and row["date_last_updated"] is not None
        fill     = _fill(DATE_CLR if has_date else NONE_CLR)
        vals = [
            row["name"],
            row["state"],
            row["site_type"],
            row["aw_active"],
            row["endpoint"],
            row["url_used"],
            row["date_last_updated"] if has_date else "—",
            int(row["asset_count_1yr"]) if pd.notna(row["asset_count_1yr"]) else 0,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = _font()
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = _border()
            if c == 7 and has_date:
                cell.number_format = "YYYY-MM-DD"

    for c, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(path)


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input",   default=None, help="Path to legistar_scrape_results_*.xlsx")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS)
    args = parser.parse_args()
    run(args.input, args.workers)
