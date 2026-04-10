"""
verify_civic_scraper.py
-----------------------
Verifies that each government website row in AW_civic_scraper_sites.xlsx
can be scraped using civic-scraper, then writes TRUE/FALSE boolean values
("checkboxes") to the `aw_active` column.

Two modes:
  1. static_preflight()  – URL-structure analysis, no network needed.
                           Runs right now, gives an immediate first pass.
  2. run_verification()  – Live HTTP checks (concurrent).
                           Run this on your local machine where outbound
                           HTTPS is available.  It overwrites the static
                           results with ground-truth values.

Quick start (local machine):
    pip install requests openpyxl pandas
    python verify_civic_scraper.py

Optional (NOT called automatically) – actual scraping via civic-scraper:
    See scrape_civicplus(), scrape_legistar(), scrape_all_sites() at the bottom.
    Requires:  pip install civic-scraper
"""

import io
import re
import time
import urllib.parse
import zipfile
import urllib3
import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_PATH   = "Copy of AW_civic_scraper_sites.xlsx"
SHEET_NAME   = "Sheet1"
MAX_WORKERS  = 30       # concurrent HTTP checks
TIMEOUT      = 12       # seconds per request

# ── civic-scraper verification config ─────────────────────────────────────────
SCRAPER_LOOKBACK_DAYS = 365          # how far back to look for documents
SCRAPER_OUTPUT_DIR    = "./scraped_output/"  # where to save PDFs when download=True
SCRAPER_MAX_WORKERS   = 10           # lower than HTTP – each call is heavier

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# HTML fingerprints that confirm a live, platform-compatible page
CIVICPLUS_FINGERPRINTS = [
    "agendacenter",
    "civicplus",
    "agenda center",
    "meeting agendas",
    "agendas & minutes",
    "agendas and minutes",
]
LEGISTAR_FINGERPRINTS = [
    "legistar",
    "granicus",
    "calendar.aspx",
]

# Patterns in the `bugs` column that signal a known-dead site
DEAD_BUG_PATTERNS = re.compile(
    r"no.?longer|missing.?site|not.?active|moved|defunct|removed|dead",
    re.IGNORECASE,
)


# ── Static preflight (URL-structure only, no network) ────────────────────────

def _static_verdict(row: pd.Series) -> bool:
    """
    Fast heuristic based purely on URL structure and existing annotations.

    Rules (applied in order):
      1. Existing 'no'  annotation   → False
      2. Known-dead `bugs` text      → False
      3. URL path ends with /Error.aspx → False
      4. URL has no recognisable path → False
      5. CivicPlus: URL path ends with /AgendaCenter (any case) → True
      6. Legistar:  URL contains legistar.com + /Calendar.aspx  → True
      7. Fallback                    → False (unknown / unsupported)
    """
    aw     = str(row.get("aw_active", "")).strip().lower()
    bugs   = str(row.get("bugs", "")).strip()
    url    = str(row.get("url", "")).strip()
    stype  = str(row.get("site_type", "")).strip().lower()

    if aw == "no":
        return False
    if bugs and bugs != "nan" and DEAD_BUG_PATTERNS.search(bugs):
        return False

    try:
        parsed = urllib.parse.urlparse(url)
    except Exception:
        return False

    path = parsed.path.rstrip("/").lower()

    if path.endswith("/error.aspx"):
        return False
    if not parsed.netloc:
        return False

    if stype == "civicplus" and path.endswith("/agendacenter"):
        return True
    if stype == "legistar" and "legistar.com" in parsed.netloc and path.endswith("/calendar.aspx"):
        return True

    return False


def static_preflight(excel_path: str = EXCEL_PATH) -> None:
    """
    Populate `aw_active` using URL-structure analysis only.
    Safe to run without network access; gives an immediate first pass.
    """
    print(f"[static] Reading {excel_path} …")
    df = pd.read_excel(excel_path, sheet_name=SHEET_NAME)

    verdicts = {i: _static_verdict(row) for i, row in df.iterrows()}
    true_count = sum(verdicts.values())
    print(f"[static] {true_count}/{len(df)} rows marked TRUE\n")

    _write_checkboxes(excel_path, verdicts)
    print(f"[static] Saved → '{excel_path}'\n")


# ── Live HTTP verification (requires outbound HTTPS) ─────────────────────────

def _http_get(url: str) -> tuple[int, str]:
    """Return (status_code, lowercased_html). Returns (-1, '') on any error."""
    try:
        resp = requests.get(
            url, headers=HEADERS, timeout=TIMEOUT,
            allow_redirects=True, verify=False,
        )
        return resp.status_code, resp.text.lower()
    except Exception:
        return -1, ""


def verify_civicplus(url: str) -> bool:
    """Return True if the CivicPlus AgendaCenter page is reachable and valid."""
    status, html = _http_get(url)
    if status != 200:
        return False
    return any(fp in html for fp in CIVICPLUS_FINGERPRINTS)


def verify_legistar(url: str) -> bool:
    """Return True if the Legistar Calendar page is reachable and valid."""
    status, html = _http_get(url)
    if status != 200:
        return False
    return any(fp in html for fp in LEGISTAR_FINGERPRINTS)


def verify_site(row: pd.Series) -> bool:
    """Route to the correct live verifier based on site_type."""
    url       = str(row.get("url", "")).strip()
    site_type = str(row.get("site_type", "")).strip().lower()
    if not url or url == "nan":
        return False
    if site_type == "civicplus":
        return verify_civicplus(url)
    if site_type == "legistar":
        return verify_legistar(url)
    status, _ = _http_get(url)
    return status == 200


def run_verification(excel_path: str = EXCEL_PATH) -> None:
    """
    Concurrently verify every URL via live HTTP, then write TRUE/FALSE to
    `aw_active`.  Run on a machine with unrestricted HTTPS access.
    """
    print(f"[live] Reading {excel_path} …")
    df = pd.read_excel(excel_path, sheet_name=SHEET_NAME)
    n  = len(df)
    print(f"[live] {n} rows to verify\n")

    results: dict[int, bool] = {}

    def _task(idx: int, row: pd.Series) -> tuple[int, bool]:
        return idx, verify_site(row)

    start = time.time()
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(_task, i, row): i for i, row in df.iterrows()}
        done = 0
        for fut in as_completed(futures):
            idx, ok = fut.result()
            results[idx] = ok
            done += 1
            if done % 50 == 0 or done == n:
                pct    = done / n * 100
                active = sum(results.values())
                print(
                    f"  [{done:4d}/{n}]  {pct:5.1f}%  "
                    f"active so far: {active}  "
                    f"elapsed: {time.time()-start:.0f}s"
                )

    active_count = sum(results.values())
    print(f"\n[live] Done – {active_count}/{n} sites verified active "
          f"({time.time()-start:.1f}s)\n")

    _write_checkboxes(excel_path, results)
    print(f"[live] Saved → '{excel_path}'\n")


# ── civic-scraper verification (uses civic_scraper library directly) ──────────

def _date_range() -> tuple[str, str]:
    """Return (start_date, end_date) strings for the configured lookback window."""
    from datetime import date, timedelta
    end   = date.today()
    start = end - timedelta(days=SCRAPER_LOOKBACK_DAYS)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def verify_civicplus_with_scraper(url: str, download: bool = False) -> tuple[bool, int]:
    """
    Use civic_scraper's CivicPlusSite to attempt a real scrape of the URL.

    Returns (success, asset_count):
      success      – True if at least one asset was found in the lookback window
      asset_count  – number of assets returned (0 if the site failed or had none)

    Args:
        url      : CivicPlus AgendaCenter URL
        download : if True, PDFs are saved to SCRAPER_OUTPUT_DIR;
                   if False (default), only metadata is fetched — faster and
                   no files written to disk, but still proves the site is scrapable
    """
    try:
        from civic_scraper.platforms import CivicPlusSite
        import os
        start_date, end_date = _date_range()
        site   = CivicPlusSite(url)
        kwargs = {"start_date": start_date, "end_date": end_date}
        if download:
            os.makedirs(SCRAPER_OUTPUT_DIR, exist_ok=True)
            from civic_scraper.base.cache import Cache
            kwargs["cache"] = Cache(SCRAPER_OUTPUT_DIR)
        assets = site.scrape(**kwargs)
        count  = len(assets) if assets else 0
        return count > 0, count
    except Exception:
        return False, 0


def verify_legistar_with_scraper(url: str, download: bool = False) -> tuple[bool, int]:
    """
    Use civic_scraper's LegistarSite to attempt a real scrape of the URL.

    Returns (success, asset_count) — same semantics as verify_civicplus_with_scraper.
    """
    try:
        from civic_scraper.platforms import LegistarSite
        import os
        start_date, end_date = _date_range()
        site   = LegistarSite(url)
        kwargs = {"start_date": start_date, "end_date": end_date}
        if download:
            os.makedirs(SCRAPER_OUTPUT_DIR, exist_ok=True)
            from civic_scraper.base.cache import Cache
            kwargs["cache"] = Cache(SCRAPER_OUTPUT_DIR)
        assets = site.scrape(**kwargs)
        count  = len(assets) if assets else 0
        return count > 0, count
    except Exception:
        return False, 0


def _scraper_task(idx: int, row: pd.Series, download: bool) -> tuple[int, bool, int, str]:
    """Single-row task for the concurrent scraper pool.  Returns (idx, ok, count, name)."""
    url   = str(row.get("url", "")).strip()
    stype = str(row.get("site_type", "")).strip().lower()
    name  = str(row.get("name", "")).strip()
    if not url or url == "nan":
        return idx, False, 0, name
    if stype == "civicplus":
        ok, count = verify_civicplus_with_scraper(url, download=download)
    elif stype == "legistar":
        ok, count = verify_legistar_with_scraper(url, download=download)
    else:
        ok, count = False, 0
    return idx, ok, count, name


def run_scraper_verification(
    excel_path: str = EXCEL_PATH,
    download: bool = False,
) -> None:
    """
    Verify every site by running civic_scraper against it directly — the most
    thorough check possible.  Requires:  pip install civic-scraper

    For each row the function calls CivicPlusSite.scrape() (or LegistarSite.scrape()),
    looks back SCRAPER_LOOKBACK_DAYS days, and marks the row TRUE only if at
    least one agenda/minutes asset is returned.

    Args:
        excel_path : path to the spreadsheet
        download   : False (default) – fetch metadata only; proves the site works
                                       without saving any files to disk
                     True            – actually download PDFs to SCRAPER_OUTPUT_DIR;
                                       end-to-end proof that files can be retrieved

    Run:
        python verify_civic_scraper.py scraper           # metadata-only
        python verify_civic_scraper.py scraper download  # full PDF download
    """
    try:
        import civic_scraper  # noqa: F401 – check it's installed before starting
    except ImportError:
        print(
            "[scraper] ERROR: civic-scraper is not installed.\n"
            "          Run:  pip install civic-scraper\n"
            "          then re-run this script.\n"
        )
        return

    mode_label = "download=True (PDFs saved)" if download else "metadata-only (no files saved)"
    print(f"[scraper] Reading {excel_path} …")
    print(f"[scraper] Mode: {mode_label}")
    print(f"[scraper] Lookback: {SCRAPER_LOOKBACK_DAYS} days\n")

    df = pd.read_excel(excel_path, sheet_name=SHEET_NAME)
    n  = len(df)

    results: dict[int, bool] = {}
    counts:  dict[int, int]  = {}

    start = time.time()
    with ThreadPoolExecutor(max_workers=SCRAPER_MAX_WORKERS) as pool:
        futures = {
            pool.submit(_scraper_task, i, row, download): i
            for i, row in df.iterrows()
        }
        done = 0
        for fut in as_completed(futures):
            idx, ok, count, name = fut.result()
            results[idx]  = ok
            counts[idx]   = count
            done         += 1
            status_icon   = "✓" if ok else "✗"
            asset_note    = f"{count} assets" if ok else "no assets / error"
            print(f"  {status_icon} [{done:4d}/{n}]  {name[:40]:<40}  {asset_note}")

    elapsed      = time.time() - start
    active_count = sum(results.values())
    total_assets = sum(counts.values())
    print(
        f"\n[scraper] Done – {active_count}/{n} sites returned assets  "
        f"({total_assets} total assets found)  "
        f"({elapsed:.1f}s)\n"
    )
    if download:
        print(f"[scraper] Files saved to: {SCRAPER_OUTPUT_DIR}\n")

    _write_checkboxes(excel_path, results)
    print(f"[scraper] Saved → '{excel_path}'\n")

    save_scraper_results(df, results, counts, download=download)


# ── Scraper results report ────────────────────────────────────────────────────

def save_scraper_results(
    df: pd.DataFrame,
    results: dict[int, bool],
    counts: dict[int, int],
    download: bool = False,
) -> str:
    """
    Write a standalone Excel report of the scraper verification run.

    Sheet 1 "Results"  – one row per site, color-coded by outcome.
    Sheet 2 "Summary"  – totals and breakdowns by state and site_type.

    Returns the path of the saved file.
    """
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Font, PatternFill, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    today     = date.today()
    out_path  = f"scraper_results_{today}.xlsx"
    start_dt, end_dt = _date_range()

    # ── Styles ────────────────────────────────────────────────────────────────
    FONT_NAME   = "Arial"
    CLR_GREEN   = "D6F0D6"   # light green – assets found
    CLR_RED     = "FAD7D7"   # light red   – no assets / error
    CLR_HEADER  = "2E4057"   # dark navy   – header row bg
    CLR_META    = "F0F4F8"   # pale blue   – metadata rows bg

    def _font(bold=False, size=10, color="000000"):
        return Font(name=FONT_NAME, bold=bold, size=size, color=color)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border():
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)

    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=False)

    # ── Build results table ───────────────────────────────────────────────────
    KEEP_COLS = ["name", "state", "country", "gov_level", "site_type", "url"]
    out = df[KEEP_COLS].copy()
    out["assets_found"] = out.index.map(lambda i: results.get(i, False))
    out["asset_count"]  = out.index.map(lambda i: counts.get(i, 0))
    out["scrape_date"]  = str(today)
    out["lookback_days"] = SCRAPER_LOOKBACK_DAYS

    HEADERS = [
        "Name", "State", "Country", "Gov Level", "Site Type",
        "URL", "Assets Found", "Asset Count", "Scrape Date", "Lookback (days)",
    ]
    COL_WIDTHS = [28, 8, 10, 14, 12, 55, 14, 13, 13, 16]

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 1 – Results
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Results"
    ws.freeze_panes = "A5"   # freeze above data rows

    # — Metadata rows (1–3) —
    meta_labels = [
        ("Civic Scraper Verification Results", True, 12),
        (f"Run date: {today}   |   Lookback: {start_dt} → {end_dt}   |   "
         f"Mode: {'download (PDFs saved)' if download else 'metadata-only'}", False, 10),
        (f"Source: {EXCEL_PATH}", False, 9),
    ]
    for r, (text, bold, size) in enumerate(meta_labels, start=1):
        cell = ws.cell(row=r, column=1, value=text)
        cell.font      = _font(bold=bold, size=size)
        cell.fill      = _fill(CLR_META)
        cell.alignment = _left()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(HEADERS))

    # — Header row (4) —
    for c, h in enumerate(HEADERS, start=1):
        cell            = ws.cell(row=4, column=c, value=h)
        cell.font       = _font(bold=True, color="FFFFFF")
        cell.fill       = _fill(CLR_HEADER)
        cell.alignment  = _center()
        cell.border     = _border()

    ws.row_dimensions[4].height = 18

    # — Data rows (5 onward) —
    for row_i, (_, row) in enumerate(out.iterrows(), start=5):
        found = row["assets_found"]
        fill  = _fill(CLR_GREEN if found else CLR_RED)
        for col_i, val in enumerate(row.values, start=1):
            cell            = ws.cell(row=row_i, column=col_i, value=val)
            cell.font       = _font()
            cell.fill       = fill
            cell.alignment  = _left()
            cell.border     = _border()

    # Column widths
    for c, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 12

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 2 – Summary
    # ═══════════════════════════════════════════════════════════════════════════
    ss = wb.create_sheet("Summary")

    total       = len(out)
    n_true      = int(out["assets_found"].sum())
    n_false     = total - n_true
    total_assets = int(out["asset_count"].sum())

    def _sh(row, col, value, bold=False, fill_hex=None, num_fmt=None):
        cell = ss.cell(row=row, column=col, value=value)
        cell.font      = _font(bold=bold)
        cell.alignment = _left()
        if fill_hex:
            cell.fill = _fill(fill_hex)
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    # Title
    title_cell = ss.cell(row=1, column=1, value="Summary")
    title_cell.font = _font(bold=True, size=12)
    title_cell.fill = _fill(CLR_META)
    ss.merge_cells("A1:D1")

    # Top-line stats
    stats = [
        ("Total sites checked",     total,        None),
        ("Sites with assets found",  n_true,       "0"),
        ("Sites without assets",     n_false,      "0"),
        ("Pct with assets",          n_true/total if total else 0, "0.0%"),
        ("Total assets found",       total_assets, "#,##0"),
        ("Lookback window (days)",   SCRAPER_LOOKBACK_DAYS, "0"),
        ("Run date",                 str(today),   None),
    ]
    for r, (label, value, fmt) in enumerate(stats, start=3):
        _sh(r, 1, label, bold=True)
        c = _sh(r, 2, value)
        if fmt:
            c.number_format = fmt

    # By site_type
    _sh(11, 1, "By site_type", bold=True, fill_hex=CLR_META)
    _sh(11, 2, "Total",        bold=True, fill_hex=CLR_META)
    _sh(11, 3, "With assets",  bold=True, fill_hex=CLR_META)
    _sh(11, 4, "% with assets",bold=True, fill_hex=CLR_META)
    for r_off, (stype, grp) in enumerate(out.groupby("site_type"), start=12):
        n_grp  = len(grp)
        n_ok   = int(grp["assets_found"].sum())
        pct    = n_ok / n_grp if n_grp else 0
        _sh(r_off, 1, stype)
        _sh(r_off, 2, n_grp,  num_fmt="#,##0")
        _sh(r_off, 3, n_ok,   num_fmt="#,##0")
        _sh(r_off, 4, pct,    num_fmt="0.0%")

    # By state (top 20 by site count)
    state_start = 11 + 2 + out["site_type"].nunique() + 2
    _sh(state_start, 1, "By state (top 20 by volume)", bold=True, fill_hex=CLR_META)
    _sh(state_start, 2, "Total",                       bold=True, fill_hex=CLR_META)
    _sh(state_start, 3, "With assets",                 bold=True, fill_hex=CLR_META)
    _sh(state_start, 4, "% with assets",               bold=True, fill_hex=CLR_META)
    state_grps = (
        out.groupby("state")[["assets_found"]]
           .apply(lambda g: pd.Series({
               "total": len(g),
               "with_assets": int(g["assets_found"].sum()),
           }), include_groups=False)
           .sort_values("total", ascending=False)
           .head(20)
    )
    for r_off, (state, row) in enumerate(state_grps.iterrows(), start=state_start + 1):
        pct = row["with_assets"] / row["total"] if row["total"] else 0
        _sh(r_off, 1, state)
        _sh(r_off, 2, int(row["total"]),        num_fmt="#,##0")
        _sh(r_off, 3, int(row["with_assets"]),  num_fmt="#,##0")
        _sh(r_off, 4, pct,                      num_fmt="0.0%")

    for col, w in zip("ABCD", [28, 16, 14, 16]):
        ss.column_dimensions[col].width = w

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save(out_path)
    print(f"[scraper] Results report → '{out_path}'")
    return out_path


# ── Shared Excel writer ───────────────────────────────────────────────────────

def _col_letter(col_name: str, excel_path: str, sheet_name: str) -> str:
    """
    Return the Excel column letter (e.g. 'I') for a given header name.
    Uses pandas to read column positions, then converts to letter.
    """
    df = pd.read_excel(excel_path, sheet_name=sheet_name, nrows=0)
    cols = list(df.columns)
    if col_name not in cols:
        raise RuntimeError(f"Column '{col_name}' not found in sheet header.")
    idx = cols.index(col_name)   # 0-based
    # Convert 0-based index to Excel column letter (A=0, B=1, …, Z=25, AA=26, …)
    col_num = idx + 1            # 1-based
    letter  = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        letter = chr(ord("A") + remainder) + letter
    return letter


def _sheet_zip_path(excel_path: str, sheet_name: str) -> str:
    """
    Return the ZIP-internal path (e.g. 'xl/worksheets/sheet1.xml') for a
    named sheet, by parsing workbook.xml and its .rels file.
    """
    with zipfile.ZipFile(excel_path) as z:
        wb_xml   = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    # rId for the sheet name
    m = re.search(
        r'<sheet\b[^>]*\bname="' + re.escape(sheet_name) + r'"[^>]*\br:id="([^"]+)"',
        wb_xml,
    )
    if not m:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in workbook.")
    rid = m.group(1)

    # Target path for that rId
    t = re.search(r'<Relationship\b[^>]*\bId="' + re.escape(rid) + r'"[^>]*\bTarget="([^"]+)"', rels_xml)
    if not t:
        raise RuntimeError(f"Relationship '{rid}' not found in workbook.rels.")
    target = t.group(1)          # e.g. "worksheets/sheet1.xml"
    return f"xl/{target}" if not target.startswith("xl/") else target


def _write_checkboxes(excel_path: str, verdicts: dict[int, bool]) -> None:
    """
    Write TRUE/FALSE boolean values to the `aw_active` column.
    Uses direct ZIP/XML patching to work even when openpyxl can't load the file
    (e.g. when pivot-cache XML contains invalid characters).
    """
    aw_col    = _col_letter("aw_active", excel_path, SHEET_NAME)
    sheet_key = _sheet_zip_path(excel_path, SHEET_NAME)

    def make_bool_cell(row_num: int, value: bool) -> str:
        v = "1" if value else "0"
        return f'<c r="{aw_col}{row_num}" s="1" t="b"><v>{v}</v></c>'

    buf = io.BytesIO()
    with zipfile.ZipFile(excel_path, "r") as zin, \
         zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_key:
                xml      = data.decode("utf-8")
                parts    = []
                pos      = 0
                i_pat    = re.compile(
                    rf'<c r="{aw_col}(\d+)"[^>]*>.*?</c>', re.DOTALL
                )
                for row_m in re.finditer(
                    r'(<row[^>]+r="(\d+)"[^>]*>)(.*?)(</row>)', xml, re.DOTALL
                ):
                    row_num  = int(row_m.group(2))
                    if row_num == 1:
                        parts.append(xml[pos:row_m.end()])
                        pos = row_m.end()
                        continue
                    df_idx = row_num - 2
                    if df_idx not in verdicts:
                        continue
                    value     = verdicts[df_idx]
                    new_cell  = make_bool_cell(row_num, value)
                    row_inner = row_m.group(3)
                    if i_pat.search(row_inner):
                        new_inner = i_pat.sub(new_cell, row_inner)
                    else:
                        next_col  = chr(ord(aw_col) + 1)
                        insert_at = re.search(
                            r'<c r="[' + next_col + r'-Z]\d+"', row_inner
                        )
                        ip        = insert_at.start() if insert_at else len(row_inner)
                        new_inner = row_inner[:ip] + new_cell + row_inner[ip:]
                    parts.append(xml[pos:row_m.start()])
                    parts.append(row_m.group(1) + new_inner + row_m.group(4))
                    pos = row_m.end()
                parts.append(xml[pos:])
                data = "".join(parts).encode("utf-8")
            zout.writestr(item, data)

    with open(excel_path, "wb") as f:
        f.write(buf.getvalue())


# ── Optional scraping functions (NOT called automatically) ────────────────────
# Requires:  pip install civic-scraper

def scrape_civicplus(url: str, start_date: str | None = None, end_date: str | None = None):
    """
    Scrape meeting agendas/minutes from a single CivicPlus AgendaCenter URL.

    Args:
        url        : e.g. 'https://ca-eastpaloalto.civicplus.com/AgendaCenter'
        start_date : 'YYYY-MM-DD'  (optional)
        end_date   : 'YYYY-MM-DD'  (optional)

    Returns:
        AssetCollection – iterate or call .to_csv(output_dir) to save.

    Example:
        assets = scrape_civicplus(
            'https://ca-eastpaloalto.civicplus.com/AgendaCenter',
            start_date='2023-01-01',
            end_date='2023-12-31',
        )
        assets.to_csv('./output/')
    """
    from civic_scraper.platforms import CivicPlusSite  # pip install civic-scraper
    site   = CivicPlusSite(url)
    kwargs = {}
    if start_date:
        kwargs["start_date"] = start_date
    if end_date:
        kwargs["end_date"] = end_date
    return site.scrape(**kwargs)


def scrape_legistar(base_url: str):
    """
    Scrape meeting data from a Legistar Calendar URL.

    Args:
        base_url : e.g. 'https://newark.legistar.com/Calendar.aspx'

    Returns:
        AssetCollection from civic-scraper.

    Example:
        assets = scrape_legistar('https://newark.legistar.com/Calendar.aspx')
        assets.to_csv('./output/')
    """
    from civic_scraper.platforms import LegistarSite  # pip install civic-scraper
    return LegistarSite(base_url).scrape()


def scrape_all_sites(
    excel_path: str = EXCEL_PATH,
    output_dir: str = "./scraped_output/",
) -> None:
    """
    Iterate over every TRUE row in the spreadsheet and scrape its documents
    using civic-scraper.  Only rows where `aw_active` is True are scraped.

    NOT CALLED AUTOMATICALLY – invoke manually when you want a full scrape.

    Example:
        scrape_all_sites(
            excel_path='Copy of AW_civic_scraper_sites.xlsx',
            output_dir='./scraped_output/',
        )
    """
    import os
    os.makedirs(output_dir, exist_ok=True)

    df     = pd.read_excel(excel_path, sheet_name=SHEET_NAME)
    active = df[df["aw_active"] == True]  # noqa: E712
    print(f"Scraping {len(active)} active sites …")

    for _, row in active.iterrows():
        url   = str(row["url"]).strip()
        stype = str(row["site_type"]).strip().lower()
        name  = str(row["name"]).strip()
        print(f"  {name} ({stype}) → {url}")
        try:
            if stype == "civicplus":
                assets = scrape_civicplus(url)
            elif stype == "legistar":
                assets = scrape_legistar(url)
            else:
                print(f"    ⚠  Unknown site_type '{stype}', skipping.")
                continue
            assets.to_csv(output_dir)
            print(f"    ✓  {len(assets)} assets saved.")
        except Exception as exc:
            print(f"    ✗  Error: {exc}")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    mode = sys.argv[1] if len(sys.argv) > 1 else "static"

    if mode == "live":
        # HTTP reachability check – requires outbound HTTPS
        run_verification()

    elif mode == "scraper":
        # civic-scraper direct check – requires pip install civic-scraper
        # Optional second arg "download" actually saves PDFs to SCRAPER_OUTPUT_DIR
        do_download = len(sys.argv) > 2 and sys.argv[2] == "download"
        run_scraper_verification(download=do_download)

    else:
        # Default: URL-structure analysis, no network needed
        static_preflight()
