"""
Build migration_research_report.xlsx from:
  1. HTTP check results (/tmp/no_assets_http_results.json)
  2. Manual research findings from web-search agents (hardcoded below)
"""
import json
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TODAY = date.today().isoformat()
OUT   = f"Original Sites to Verify/legistar_civicplus_migration_research_{TODAY}.xlsx"

# ── 1. Load HTTP check results ───────────────────────────────────────────────
with open('/tmp/no_assets_http_results.json') as f:
    http_results = json.load(f)

http_lookup = {}
for r in http_results:
    key = (r['name'].strip().lower(), r['state'].strip().upper())
    http_lookup[key] = r

# ── 2. Manual research findings (from web-search agents) ────────────────────
# Format: name, state, new_platform, new_url, notes
MANUAL = [
    # TX
    ("Frisco",         "TX", "OnBase Agenda Online (Hyland)", "agenda.friscotexas.gov", "Old AgendaCenter still accessible as archive; current council agendas on OnBase"),
    ("Rowlett",        "TX", "CivicClerk (CivicPlus)",        "rowletttx.portal.civicclerk.com", "New city domain rowletttx.gov; historical records in Laserfiche"),
    ("Sherman",        "TX", "CivicClerk (CivicPlus)",        "shermantx.portal.civicclerk.com", "Old AgendaCenter still loads at ci.sherman.tx.us as archive"),
    ("Wichita Falls",  "TX", "Swagit (Granicus)",             "wichitafallstx.new.swagit.com", "New domain wichitafallstx.gov; agenda PDF packets in DocumentCenter"),
    ("Lewisville",     "TX", "Legistar (Granicus)",           "cityoflewisville.legistar.com", "Fully migrated; new domain cityoflewisville.com"),
    ("Denton",         "TX", "Legistar (Granicus)",           "denton-tx.legistar.com", "AgendaCenter kept for boards/archive; City Council fully on Legistar"),
    ("Euless",         "TX", "Swagit (Granicus)",             "eulesstx.new.swagit.com", "New domain eulesstx.gov; meeting videos and agendas on Swagit"),
    # CA
    ("Menifee",        "CA", "PrimeGov (Granicus)",           "cityofmenifee.primegov.com", "AgendaCenter still live but PrimeGov is primary for current council meetings"),
    ("Fremont",        "CA", "CivicEngage / CivicPlus (rebuit)", "fremont.gov/government/agenda-center", "Domain migrated to fremont.gov; rebuilt on CivicEngage under new URL structure"),
    ("Palmdale",       "CA", "Granicus",                      "palmdale.granicus.com", "Domain changed to cityofpalmdaleca.gov; Council agendas on Granicus"),
    ("Orange",         "CA", "Legistar + Granicus",           "cityoforange.legistar.com", "Dual platform: Legistar for calendar, Granicus for video; old AgendaCenter URL blocked"),
    ("La Mesa",        "CA", "eScribe (Diligent)",            "pub-lamesa.escribemeetings.com", "Domain migrated to cityoflamesa.gov; newer meetings use eScribe"),
    ("Danville",       "CA", "Granicus",                      "danville-ca.granicus.com", "AgendaCenter still active; Granicus is primary for video and agenda access"),
    # FL
    ("Bay County",     "FL", "NovusAgenda (Granicus)",        "baycountyfl.gov/311/Agendas", "Agendas via NovusAgenda embedded on county site; old civicplus.com subdomain gone"),
    ("Miramar",        "FL", "Legistar (Granicus)",           "miramar.legistar.com", "Fully migrated to Legistar; video on miramar.granicus.com"),
    ("Pensacola",      "FL", "Legistar (Granicus)",           "pensacola.legistar.com", "AgendaCenter kept as pre-2020 archive; all current boards on Legistar"),
    ("Fort Myers",     "FL", "Granicus",                      "cityftmyers.granicus.com", "Domain changed to fortmyers.gov; City Council fully on Granicus"),
    ("Pasco County",   "FL", "CivicClerk (CivicPlus)",        "pascocofl.portal.civicclerk.com", "Domain migrated to pascocountyfl.gov; AgendaCenter path 404 on new domain"),
    ("New Smyrna Beach","FL","Granicus",                      "cityofnsb.granicus.com", "AgendaCenter still loads but links out to Granicus as the primary platform"),
    # CO
    ("Pueblo",         "CO", "CivicClerk (CivicPlus)",        "puebloco.portal.civicclerk.com", "Still in CivicPlus ecosystem; old AgendaCenter kept as archive"),
    ("Summit County",  "CO", "Diligent OneFile",              "summitcountyco.community.diligentoneplatform.com/Portal", "Never used standard CivicPlus AgendaCenter; website on Revize CMS"),
    ("Glenwood Springs","CO","CivicEngage / CivicPlus (domain change)", "cogs.us/AgendaCenter", "Domain changed from gwsco.gov to cogs.us; AgendaCenter fully operational"),
    ("Aspen",          "CO", "CivicEngage / CivicPlus (domain change)", "aspen.gov/agendacenter", "Domain migrated to aspen.gov; AgendaCenter live and current"),
    ("Arapahoe County","CO", "Legistar (Granicus)",           "arapahoe.legistar.com", "Fully migrated off CivicPlus; new domain arapahoeco.gov"),
    ("Jefferson County","CO","eSCRIBE (Diligent)",            "pub-jeffco.escribemeetings.com", "CivicPlus AgendaCenter officially retired Dec 2022; pre-Oct 2022 video on Granicus"),
    # KS
    ("Merriam",        "KS", "CivicEngage / CivicPlus (still active)", "merriam.org/AgendaCenter", "403 may be bot-blocking; city still on CivicPlus platform"),
    ("Overland Park",  "KS", "CivicWeb (Diligent/iCompass)",  "opkansas.civicweb.net/Portal/MeetingTypeList.aspx", "Migrated to CivicWeb; new city domain opkansas.org"),
    ("Olathe",         "KS", "Legistar (Granicus)",           "olatheks.legistar.com", "Fully migrated to Legistar; new domain olatheks.gov"),
    ("Lenexa",         "KS", "Self-hosted PDFs",              "lenexa.com/Government/Agendas-Minutes", "Moved entirely to self-hosted PDFs; no third-party agenda platform"),
    # MN
    ("Brooklyn Park",  "MN", "Self-hosted (WordPress) + Granicus video", "brooklynpark.org/city-council/city-council-documents/", "PDFs on WordPress; video archive on Granicus"),
    ("Woodbury",       "MN", "CivicEngage / CivicPlus (custom domain)", "woodburymn.gov/AgendaCenter", "Still on CivicPlus; moved from civicplus.com subdomain to own domain"),
    ("Plymouth",       "MN", "CivicClerk (CivicPlus)",        "plymouthmn.portal.civicclerk.com", "Still in CivicPlus ecosystem; migrated from old subdomain to CivicClerk"),
    ("Eagan",          "MN", "Self-hosted + Laserfiche archive", "cityofeagan.com/meetings", "Current agendas on city site; archive in Laserfiche at eagandocs.cityofeagan.com"),
    # OH
    ("Dublin",         "OH", "Self-hosted (WordPress)",       "dublinohiousa.gov/council/meeting-schedule/", "Documents as PDFs on dublinohiousa.gov; no third-party agenda platform"),
    ("Hilliard",       "OH", "CivicWeb (Diligent/iCompass)",  "hilliardohio.civicweb.net/Portal/", "Migrated to CivicWeb; new domain hilliardohio.gov"),
    ("Westerville",    "OH", "Self-hosted",                   "westerville.org/government/clerk-of-council/meeting-agendas-and-minutes", "Documents organized by year on city website"),
    ("Medina County",  "OH", "Self-hosted (WordPress)",       "medinaco.org/county-commissioners/", "Agendas and minutes as PDFs on county WordPress site"),
    # PA
    ("Upper Darby",    "PA", "EvoGov",                        "upperdarby.org/meetingdashboard", "Migrated to EvoGov; PDFs hosted on evocloud S3"),
    ("Lower Merion",   "PA", "Granicus + NovusAgenda",        "lowermerion.granicus.com", "Dual platform: Granicus for video, NovusAgenda for agenda packets"),
    ("Bucks County",   "PA", "CivicClerk (CivicPlus)",        "buckscopa.portal.civicclerk.com", "CivicClerk is new primary; old AgendaCenter still accessible"),
    ("Chester County", "PA", "CivicClerk (CivicPlus)",        "chestercopa.portal.civicclerk.com", "Same dual-platform as Bucks County; newer docs going to CivicClerk"),
    # WI
    ("Eau Claire",     "WI", "Granicus",                      "eauclairewi.granicus.com", "Granicus for agenda viewer, video, PDFs; city CMS still CivicEngage"),
    ("Kenosha",        "WI", "Granicus (video/archive) + own website", "kenosha.org/agendas-and-minutes", "Granicus for archived video; agendas hosted on kenosha.org"),
    ("Sheboygan",      "WI", "Municode Meetings (CivicPlus)", "sheboygan-wi.municodemeetings.com", "Municode is a CivicPlus subsidiary; AgendaCenter page still loads but directs to Municode"),
    # WA
    ("Renton",         "WA", "CivicWeb (Diligent/iCompass)",  "renton.civicweb.net", "CivicWeb is Diligent-branded; records back to 2010; not related to CivicPlus"),
    ("Kirkland",       "WA", "PrimeGov (Granicus)",           "kirklandwa.primegov.com", "PrimeGov acquired by Granicus via Rock Solid Technologies"),
    ("Kennewick",      "WA", "CivicClerk (CivicPlus)",        "kennewickwa.portal.civicclerk.com", "Old AgendaCenter loads but has 'we have moved' banner directing to CivicClerk"),
    ("Spokane Valley", "WA", "QScend + Laserfiche archive",   "spokanevalleywa.qscend.com/agendas", "QScend for current; Laserfiche for archive; CivicEngage still runs main site"),
    # MO
    ("Lee's Summit",   "MO", "Legistar (Granicus)",           "lsmo.legistar.com", "Fully migrated to Legistar; Granicus video at lsmo.granicus.com"),
    ("O'Fallon",       "MO", "Custom/self-hosted",            "agenda.ofallon.mo.us", "Own agenda subdomain; no major third-party platform"),
    ("Independence",   "MO", "CivicClerk (CivicPlus)",        "independencemo.portal.civicclerk.com", "Pre-April 2024 materials archived on NovusAgenda; CivicClerk is current"),
    # MA
    ("Waltham",        "MA", "CivicEngage / CivicPlus (own domain)", "city.waltham.ma.us/minutes-and-agendas", "Agendas organized by year on city website; still on CivicPlus CMS"),
    ("Quincy",         "MA", "Revize (own website)",          "quincyma.gov/government/elected_officials/city_council/council_agendas___minutes/", "Moved to Revize CMS; all agendas hosted natively"),
    ("Brockton",       "MA", "Self-hosted (WordPress)",       "brockton.ma.us/public-documents/", "Agendas and minutes in document library on WordPress site"),
]

manual_lookup = {(n.strip().lower(), s.strip().upper()): (p, u, nt) for n,s,p,u,nt in MANUAL}

# ── 3. Build full dataframe from HTTP results ────────────────────────────────
rows = []
for r in http_results:
    name  = r['name'].strip()
    state = r['state'].strip().upper()
    key   = (name.lower(), state)

    # Determine http_category
    status = str(r.get('http_status',''))
    if 'ERROR' in status or status in ['None','']:
        http_cat = 'dead/error'
    elif r.get('redirected'):
        http_cat = 'redirected'
    elif status == '200':
        http_cat = '200 (loads)'
    else:
        http_cat = f'HTTP {status}'

    # Pull manual research if available
    man = manual_lookup.get(key)
    new_platform = man[0] if man else (r.get('platform_detected') or '')
    new_url      = man[1] if man else (r.get('redirect_domain') or '')
    notes        = man[2] if man else r.get('notes','')

    rows.append({
        'name':           name,
        'state':          state,
        'gov_level':      r.get('gov_level',''),
        'original_url':   r.get('original_url',''),
        'http_status':    r.get('http_status',''),
        'http_category':  http_cat,
        'redirected':     'yes' if r.get('redirected') else 'no',
        'final_url':      r.get('final_url','') or '',
        'new_platform':   new_platform,
        'new_url':        new_url,
        'manually_researched': 'yes' if man else 'no',
        'notes':          notes,
    })

df = pd.DataFrame(rows)
df = df.sort_values(['state','name']).reset_index(drop=True)

# ── 4. Platform summary ───────────────────────────────────────────────────────
platform_counts = {}
for r in rows:
    p = r['new_platform']
    if p:
        # Normalise
        if 'granicus' in p.lower() or 'legistar' in p.lower() or 'swagit' in p.lower() or 'primegov' in p.lower() or 'novus' in p.lower():
            bucket = 'Granicus ecosystem (Legistar / PrimeGov / Swagit / NovusAgenda)'
        elif 'civicclerk' in p.lower() or 'municode' in p.lower() or 'civicengage' in p.lower() or 'civicplus' in p.lower():
            bucket = 'CivicPlus ecosystem (CivicClerk / Municode / CivicEngage)'
        elif 'diligent' in p.lower() or 'escribe' in p.lower() or 'civicweb' in p.lower() or 'icompass' in p.lower():
            bucket = 'Diligent ecosystem (eSCRIBE / CivicWeb / OneFile)'
        elif 'self' in p.lower() or 'wordpress' in p.lower() or 'own' in p.lower() or 'custom' in p.lower() or 'pdf' in p.lower():
            bucket = 'Self-hosted / own website'
        elif 'laserfiche' in p.lower():
            bucket = 'Laserfiche'
        elif 'wordpress' in p.lower():
            bucket = 'Self-hosted / own website'
        else:
            bucket = p
        platform_counts[bucket] = platform_counts.get(bucket, 0) + 1

# HTTP category summary
http_counts = df['http_category'].value_counts().to_dict()

# ── 5. Write Excel ─────────────────────────────────────────────────────────────
FONT    = "Arial"
HDR_CLR = "2E4057"

def _font(bold=False, size=10, color="000000"):
    return Font(name=FONT, bold=bold, size=size, color=color)
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)
def _border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)
def _write_header(ws, headers, widths, title_text):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1, value=title_text)
    t.font = _font(bold=True, size=11); t.fill = _fill("F0F4F8")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = _font(bold=True, color="FFFFFF"); cell.fill = _fill(HDR_CLR)
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = _border()
    ws.row_dimensions[2].height = 16
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

# Row fill colours by category
CAT_COLORS = {
    'redirected':   'FFF9C4',   # yellow
    '200 (loads)':  'E8F5E9',   # light green
    'dead/error':   'FFEBEE',   # light red
}

wb = Workbook()

# — Sheet 1: Full results —
ws1 = wb.active; ws1.title = "All No-Asset Sites"
HEADERS1 = ["Name","State","Gov Level","Original URL","HTTP Status","HTTP Category",
            "Redirected?","Final URL","New Platform (auto/researched)","New URL","Manually Researched","Notes"]
WIDTHS1  = [28,7,14,52,12,16,11,52,38,45,18,55]
_write_header(ws1, HEADERS1, WIDTHS1, f"CivicPlus No-Asset Sites — Migration Research  ·  {TODAY}  ·  517 US sites")
ws1.freeze_panes = "A3"

for r_idx, (_, row) in enumerate(df.iterrows(), start=3):
    cat = row['http_category']
    clr = CAT_COLORS.get(cat, 'FFFFFF')
    fill = _fill(clr)
    vals = [row['name'], row['state'], row['gov_level'], row['original_url'],
            str(row['http_status']), row['http_category'], row['redirected'],
            row['final_url'], row['new_platform'], row['new_url'],
            row['manually_researched'], row['notes']]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=r_idx, column=c, value=str(v) if v is not None else '')
        cell.font = _font(); cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(c in (4,8,12)))
        cell.border = _border()

# — Sheet 2: Manually researched sites —
ws2 = wb.create_sheet("Researched Sites")
HEADERS2 = ["Name","State","Gov Level","Original URL","HTTP Category","New Platform","New URL","Notes"]
WIDTHS2  = [28,7,14,52,16,38,48,60]
_write_header(ws2, HEADERS2, WIDTHS2, f"Manually Researched Sites ({len(MANUAL)} agencies)  ·  {TODAY}")
ws2.freeze_panes = "A3"
researched = df[df['manually_researched']=='yes'].reset_index(drop=True)
for r_idx, (_, row) in enumerate(researched.iterrows(), start=3):
    clr = CAT_COLORS.get(row['http_category'], 'FFFFFF')
    fill = _fill(clr)
    vals = [row['name'], row['state'], row['gov_level'], row['original_url'],
            row['http_category'], row['new_platform'], row['new_url'], row['notes']]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r_idx, column=c, value=str(v) if v is not None else '')
        cell.font = _font(); cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(c in (4,7,8)))
        cell.border = _border()

# — Sheet 3: Summary —
ws3 = wb.create_sheet("Summary")
ws3.column_dimensions['A'].width = 50
ws3.column_dimensions['B'].width = 12

def _sh(row, col, val, bold=False):
    cell = ws3.cell(row=row, column=col, value=val)
    cell.font = _font(bold=bold); cell.alignment = Alignment(horizontal="left"); cell.border = _border()

_sh(1,1,f"Migration Research Summary — {TODAY}  (517 US no-asset CivicPlus sites)", bold=True)

r = 3
_sh(r,1,"HTTP STATUS BREAKDOWN", bold=True); _sh(r,2,"Count", bold=True); r+=1
for cat, cnt in sorted(http_counts.items(), key=lambda x:-x[1]):
    _sh(r,1,cat); _sh(r,2,cnt); r+=1

r+=1
_sh(r,1,"NEW PLATFORM DETECTED (auto-detected from redirects, 172 sites)", bold=True); _sh(r,2,"Count", bold=True); r+=1
for p, c in sorted(platform_counts.items(), key=lambda x:-x[1]):
    _sh(r,1,p); _sh(r,2,c); r+=1

r+=1
_sh(r,1,"KEY FINDINGS FROM MANUAL RESEARCH (54 sites across 13 states)", bold=True); r+=1
findings = [
    "1. The dominant migration destination is the Granicus ecosystem — Legistar, Swagit, PrimeGov, NovusAgenda, and Granicus itself. Many mid-to-large cities moved here.",
    "2. CivicPlus is retaining clients but upgrading them: old *.civicplus.com subdomains are being decommissioned in favour of city-owned domains still running CivicEngage or the newer CivicClerk product.",
    "3. CivicClerk (portal.civicclerk.com) is CivicPlus's successor to AgendaCenter for agenda/minutes management. Used by: Rowlett TX, Sherman TX, Pasco County FL, Pueblo CO, Plymouth MN, Bucks/Chester County PA, Kennewick WA, Independence MO.",
    "4. Municode (municodemeetings.com) is also a CivicPlus subsidiary — so Sheboygan WI and others on Municode are still in the CivicPlus family.",
    "5. Diligent/iCompass ecosystem captures a smaller share: CivicWeb (Renton WA, Hilliard OH, Overland Park KS), eSCRIBE (Jefferson County CO, La Mesa CA), Diligent OneFile (Summit County CO).",
    "6. Domain migrations account for many '0 assets' cases — the scraper was pointed at the old domain (e.g., cityoflamesa.us) while the city moved to a new domain (cityoflamesa.gov) where AgendaCenter is still functional.",
    "7. A meaningful minority (~15%) self-host: PDFs directly on WordPress/Revize sites (Dublin OH, Brockton MA, Medina County OH, Lenexa KS), sometimes supplemented by Laserfiche for archive (Eagan MN, Spokane Valley WA).",
    "8. Some agencies kept AgendaCenter alive as a historical archive while routing all new documents to a new platform (Pensacola FL, Danville CA, Kennewick WA, Bucks County PA).",
]
for f in findings:
    _sh(r,1,f); r+=1

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"\nHTTP breakdown: {http_counts}")
print(f"Platform buckets: {platform_counts}")
